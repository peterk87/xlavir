"""Excel spreadsheet IO functions"""
import logging
from copy import copy
from pathlib import Path
from typing import List, Optional, Set, Tuple, Dict, Union

import openpyxl
import pandas as pd
from xlsxwriter.workbook import Workbook
from xlsxwriter.worksheet import Worksheet

from xlavir.images import SheetImage
from xlavir.io.excel_sheet_dataframe import ExcelSheetDataFrame, SheetName
from xlavir.qc import QualityRequirements
from xlavir.util import get_col_widths, get_row_heights
from xlavir import __version__

logger = logging.getLogger(__name__)


def copy_spreadsheet(src_path: Path,
                     dest_path: Path,
                     source_sheet_index: int = 0) -> None:
    """Copy a spreadsheet from a source Excel spreadsheet to a destination spreadsheet.

    All cell values and styles are copied over as well as row heights, column widths, merged cells.

    By default the 1st worksheet is copied over.

    Args:
        src_path (Path): Source Excel spreadsheet path
        dest_path (Path): Destination Excel spreadsheet path
        source_sheet_index (int): Source spreadsheet worksheet index to copy to destination spreadsheet
    """
    from openpyxl.cell.cell import Cell

    src_book = openpyxl.load_workbook(src_path)
    dest_book = openpyxl.load_workbook(dest_path)
    sheet = src_book.worksheets[source_sheet_index]
    new_sheet = dest_book.create_sheet(sheet.title)
    dest_book.move_sheet(new_sheet, offset=(len(dest_book.sheetnames) * -2))
    for k, v in sheet.column_dimensions.items():
        new_sheet.column_dimensions[k] = copy(v)
    new_sheet.merged_cells = copy(sheet.merged_cells)
    for k, v in sheet.row_dimensions.items():
        new_sheet.row_dimensions[k] = copy(v)
    row: Tuple[Cell]
    for row in sheet.rows:
        for cell in row:
            new_cell: Cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            new_cell.comment = copy(cell.comment)
    dest_book.save(filename=dest_path)


def write_xlsx_report(dfs: List[ExcelSheetDataFrame],
                      output_xlsx: Path,
                      quality_reqs: QualityRequirements,
                      images_for_sheets: Optional[List[SheetImage]] = None):
    """Write the output Excel XLSX file


    """
    with pd.ExcelWriter(output_xlsx, engine='xlsxwriter') as writer:
        monospace = dict(font_name='Courier New')
        text_wrap = dict(text_wrap=True)
        float_1dec = dict(num_format='0.0')
        perc_1dec = dict(num_format='0.0%')
        perc_2dec = dict(num_format='0.00%')

        book: Workbook = writer.book
        monospace_fmt = book.add_format(monospace)
        monospace_wrap_fmt = book.add_format({**monospace, **text_wrap})
        monospace_wrap_fmt.set_align('left')
        monospace_wrap_fmt.set_align('top')
        monospace_float_fmt = book.add_format({**monospace, **float_1dec})
        monospace_perc_fmt = book.add_format({**monospace, **perc_1dec})
        monospace_perc_2dec_fmt = book.add_format({**monospace, **perc_2dec})
        header_with_wrap_fmt = book.add_format({**text_wrap,
                                                'bold': True,
                                                'border': 1,
                                                'align': 'center',
                                                'valign': 'top'})
        varmap_header_fmt = book.add_format(dict(border=1,
                                                 align='left',
                                                 valign='bottom',
                                                 rotation=45,
                                                 font_name='Courier New'))
        fail_qc_fmt = book.add_format(dict(bg_color='FC9295',
                                           font_name='Courier New',
                                           bold=True))
        pass_qc_fmt = book.add_format(dict(bg_color='c4edce',
                                           font_name='Courier New',
                                           bold=False))
        float_cols = {'Mean Coverage Depth', 'Mean Depth'}
        perc_cols = {'% Genome Coverage'}
        perc_2dec_cols = {'Alternate Allele Frequency', 'Min AF', 'Max AF', 'Mean AF'}

        images_added = False

        for esdf in dfs:
            if images_for_sheets and esdf.sheet_name == SheetName.workflow_info.value:
                add_images(images_for_sheets, book)
                images_added = True
            esdf.df.to_excel(writer, sheet_name=esdf.sheet_name, **esdf.pd_to_excel_kwargs)

            sheet: Worksheet = book.get_worksheet_by_name(esdf.sheet_name)

            idx_and_cols = [esdf.df.index.name] + list(esdf.df.columns)

            if esdf.header_comments:
                for i, col_name in enumerate(idx_and_cols):
                    if col_name in esdf.header_comments:
                        sheet.write_comment(0, i, esdf.header_comments[col_name])

            if esdf.autofit:
                for i, (width, col_name) in enumerate(zip(get_col_widths(esdf.df,
                                                                         index=True,
                                                                         max_width=80,
                                                                         include_header=esdf.include_header_width),
                                                          idx_and_cols)):
                    if col_name in float_cols:
                        sheet.set_column(i, i, width, monospace_float_fmt)
                    elif col_name in perc_cols:
                        sheet.set_column(i, i, width, monospace_perc_fmt)
                    elif col_name in perc_2dec_cols:
                        sheet.set_column(i, i, width, monospace_perc_2dec_fmt)
                    else:
                        sheet.set_column(i, i, width, monospace_fmt)
                if not esdf.include_header_width:
                    for i, col_name in enumerate(idx_and_cols):
                        sheet.write_string(0, i, string=col_name, cell_format=header_with_wrap_fmt)

            elif esdf.column_widths:
                for i, (width, col_name) in enumerate(zip(esdf.column_widths, idx_and_cols)):
                    sheet.set_column(i, i, width, monospace_wrap_fmt)

                for i, idx in enumerate(esdf.df.index, 1):
                    sheet.set_row(i, get_row_heights(esdf.df, idx), monospace_wrap_fmt)

            if esdf.sheet_name == SheetName.varmat.value:
                sheet.write_comment(row=0,
                                    col=0,
                                    comment=f'This sheet contains a matrix of alternate allele variant observation'
                                            f' frequency values for samples and variants. '
                                            f'3-colour conditional formatting is applied to the variant '
                                            f'frequency values where a major variant '
                                            f'(e.g. alternate allele frequency >={quality_reqs.major_allele_freq}) '
                                            f'is highlighted in green. Red indicates where the allele variant is not '
                                            f'observed in the sample (e.g. alternate allele frequency equals 0.0).')
                sheet.set_row(0, max(len(x) for x in idx_and_cols) * 5)
                for i, col_name in enumerate(idx_and_cols):
                    if i == 0:
                        continue
                    sheet.write_string(0, i, string=col_name, cell_format=varmap_header_fmt)
                sheet.conditional_format(first_row=1,
                                         first_col=1,
                                         last_row=esdf.df.shape[0],
                                         last_col=esdf.df.shape[1],
                                         options=dict(type='3_color_scale',
                                                      min_type='num',
                                                      mid_type='num',
                                                      max_type='num',
                                                      min_value=0.0,
                                                      mid_value=quality_reqs.major_allele_freq,
                                                      max_value=1.0))

            if esdf.sheet_name == SheetName.consensus.value:
                sheet.hide_gridlines(2)
                sheet.hide_row_col_headers()

            if esdf.sheet_name == SheetName.qc_stats.value:
                columns = esdf.df.columns.tolist()
                qc_status_column = 'QC Status'
                column_idx = columns.index(qc_status_column) + 1
                logger.info(f'Setting red background color for "{qc_status_column}" column (index={column_idx}) '
                            f'of sheet "{esdf.sheet_name}"')
                sheet.conditional_format(1, column_idx, esdf.df.shape[0], column_idx, options=dict(type='cell',
                                                                                                   value='"FAIL"',
                                                                                                   criteria='equal to',
                                                                                                   format=fail_qc_fmt))
                sheet.conditional_format(1, column_idx, esdf.df.shape[0], column_idx, options=dict(type='cell',
                                                                                                   value='"PASS"',
                                                                                                   criteria='equal to',
                                                                                                   format=pass_qc_fmt))
                cond_fmt_3color = dict(type='3_color_scale',
                                       min_type='num',
                                       mid_type='num',
                                       max_type='num', )

                add_cond_fmt(sheet,
                             esdf.df,
                             '% Genome Coverage',
                             {**cond_fmt_3color, **dict(min_value=0.0,
                                                        mid_value=quality_reqs.min_genome_coverage,
                                                        max_value=1.0)})
                for column in ['Median Coverage Depth', 'Mean Coverage Depth']:
                    add_cond_fmt(sheet,
                                 esdf.df,
                                 column,
                                 {**cond_fmt_3color, **dict(min_value=0,
                                                            mid_value=quality_reqs.min_median_depth,
                                                            max_type='max')}
                                 )
                ref_len = esdf.df['Reference Sequence Length'].values[0]
                for column in esdf.df.columns[esdf.df.columns.str.contains('X positions')]:
                    add_cond_fmt(sheet,
                                 esdf.df,
                                 column,
                                 {
                                     **cond_fmt_3color,
                                     **dict(min_color='6eb758',
                                            max_color='930e11',
                                            mid_color='f73639',
                                            min_value=0,
                                            mid_value=ref_len * 0.05,
                                            max_value=ref_len)}
                                 )
                for column in esdf.df.columns[esdf.df.columns.str.contains('Reads')]:
                    add_cond_fmt(sheet,
                                 esdf.df,
                                 column,
                                 {
                                     **cond_fmt_3color,
                                     **dict(min_color='930e11',
                                            mid_color='f4cf46',
                                            max_color='6eb758',
                                            min_value=0,
                                            mid_value=1000,
                                            max_type='max')}
                                 )

        if images_for_sheets and not images_added:
            add_images(images_for_sheets, book)

    df_qc = get_qc_df(dfs)
    failed_samples = set(df_qc[df_qc['QC Status'] == 'FAIL'].index)
    add_comments(xlsx_path=output_xlsx, failed_samples=failed_samples, esdfs=dfs)


def add_cond_fmt(sheet: Worksheet,
                 df: pd.DataFrame,
                 column: str,
                 options: dict):
    columns = df.columns.tolist()
    column_idx = columns.index(column) + 1
    logger.info(f'Applying conditional formatting to column "{column}" (index={column_idx}) of '
                f'sheet "{sheet.name}"')

    sheet.conditional_format(first_row=1,
                             first_col=column_idx,
                             last_row=df.shape[0],
                             last_col=column_idx,
                             options=options)


def get_qc_df(dfs: List[ExcelSheetDataFrame]) -> Optional[pd.DataFrame]:
    for esdf in dfs:
        if esdf.sheet_name == SheetName.qc_stats.value:
            return esdf.df


def add_images(images_for_sheets: List[SheetImage],
               book: Workbook):
    """Add images and their descriptions to new sheets in a workbook"""
    import imageio
    text_wrap_fmt = book.add_format(dict(text_wrap=True, valign='justify'))
    for sheet_image in images_for_sheets:
        sheet = book.add_worksheet(sheet_image.sheet_name)
        sheet.set_column(0, 0, 100, text_wrap_fmt)
        sheet.write(0, 0, sheet_image.image_description, text_wrap_fmt)
        img = imageio.imread(sheet_image.image_path)
        x_size, y_size, _ = img.shape
        yx_ratio = y_size / x_size
        logger.debug(f'Image "{sheet_image.image_path.name}", x={x_size}, y={y_size}, y/x={yx_ratio}')
        sheet.insert_image(1, 0, sheet_image.image_path, options=dict(x_scale=1.0,
                                                                      y_scale=yx_ratio,
                                                                      object_position=3))
        sheet.hide_gridlines(2)
        sheet.hide_row_col_headers()


def get_excel_sheet_df(esds: List[ExcelSheetDataFrame],
                       sheet_name: str) -> Optional[ExcelSheetDataFrame]:
    for esd in esds:
        if esd.sheet_name == sheet_name:
            return esd


def add_comments(xlsx_path: Path,
                 failed_samples: Set[str],
                 esdfs: List[ExcelSheetDataFrame]) -> None:
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.dimensions import ColumnDimension
    logger.info(f'Loading workbook "{xlsx_path.name}" with openpyxl '
                f'to highlight {len(failed_samples)} samples that have failed QC')
    book = openpyxl.load_workbook(xlsx_path)
    logger.info(f'Loaded "{xlsx_path.name}" using openpyxl. Sheets: {book.get_sheet_names()}')
    logger.info(f'Adjusting comment textbox sizes to fit text')
    for sheetname in book.sheetnames:
        for row in book[sheetname]:
            for cell in row:
                if cell.comment:
                    comment: Comment = cell.comment
                    comment.width = 300
                    comment.height = max((100, len(comment.text) / 3 * 2))
                    comment.author = f'xlavir version {__version__}'

    sheet_names = [
        SheetName.pangolin.value,
        SheetName.variants.value,
        SheetName.varmat.value,
    ]
    light_red = 'FC9295'
    for sheet_name in sheet_names:
        try:
            sheet: Worksheet = book[sheet_name]
            logger.info(f'Highlighting failed samples in sheet "{sheet_name}".')
            for i, row in enumerate(sheet.rows):
                if i == 0:
                    continue
                cell = row[0]
                if cell.value in failed_samples:
                    cell.comment = Comment(f'Warning: Sample "{cell.value}" has failed general NGS QC',
                                           author='xlavir')
                    cell.fill = PatternFill(fill_type='solid',
                                            fgColor=light_red)
        except KeyError:
            pass

    esd_varmat = get_excel_sheet_df(esdfs, SheetName.varmat.value)
    esd_variants = get_excel_sheet_df(esdfs, SheetName.variants.value)
    if esd_varmat and esd_variants:
        try:
            sheet: Worksheet = book[SheetName.varmat.value]
            logger.info(f'Adding additional comments to variant matrix values')
            df_varmat = esd_varmat.df
            variants: Dict[Tuple[str, str], Dict[str, Union[str, float, int]]] = esd_variants \
                .df.reset_index() \
                .set_index(['Sample', 'Mutation']) \
                .to_dict(orient='index')

            for i, row in enumerate(sheet.rows):
                if i == 0:
                    continue
                sample = df_varmat.index.values[i - 1]
                for j, cell in enumerate(row):
                    if j == 0:
                        continue
                    mutation = df_varmat.columns.values[j - 1]
                    variant = variants.get((sample, mutation), None)
                    if variant:
                        variant_str = '\n'.join(f'{k}: {v}' for k, v in variant.items())
                        comment_text = f'Sample: {sample}\nMutation: {mutation}\n{variant_str}'
                    else:
                        comment_text = f'Mutation "{mutation}" not found in sample "{sample}"'
                    cell.comment = Comment(comment_text,
                                           author=f'xlavir version {__version__}',
                                           width=300,
                                           height=len(comment_text))
        except KeyError:
            pass

    try:
        sheet: Worksheet = book[SheetName.consensus.value]

        sheet.column_dimensions['A'] = ColumnDimension(worksheet=sheet,
                                                       index='A',
                                                       width=100)

        logger.info(f'Highlighting consensus sequences of failed '
                    f'samples in sheet "{SheetName.consensus.value}".')
        highlight_seq = False
        sample_name = ''

        dark_red = '260000'
        for i, row in enumerate(sheet.rows):
            cell = row[0]
            if cell.value[0] == '>':
                sample_name = cell.value[1:]
                if sample_name in failed_samples:
                    highlight_seq = True
                    cell.comment = Comment(f'Warning: Sample "{sample_name}" has failed general NGS QC',
                                           author='xlavir')
                    cell.fill = PatternFill(fill_type='solid', fgColor=light_red)
                    font: Font = cell.font
                    cell.font = Font(name='Courier New',
                                     color=dark_red,
                                     size=font.size,
                                     family=font.family)
                else:
                    font: Font = cell.font
                    cell.font = Font(name='Courier New',
                                     color='000000',
                                     size=font.size,
                                     family=font.family)
                    highlight_seq = False
            elif cell.value and highlight_seq:
                cell.comment = Comment(f'Warning: Sample "{sample_name}" has failed general NGS QC',
                                       author='xlavir')

                cell.fill = PatternFill(fill_type='solid', fgColor=light_red)
                font: Font = cell.font
                cell.font = Font(name='Courier New',
                                 color=dark_red,
                                 size=font.size,
                                 family=font.family)
                highlight_seq = False
            else:
                font: Font = cell.font
                cell.font = Font(name='Courier New',
                                 color='000000',
                                 size=font.size,
                                 family=font.family)
    except KeyError:
        pass
    book.save(xlsx_path)
